import openpyxl


workbook = openpyxl.load_workbook("Assignment_Timecard.xlsx")

sheets = workbook.sheetnames
Sheet1= workbook['Sheet1']


sheet_obj = workbook.active


row = sheet_obj.max_row     # find the no of rows
column = sheet_obj.max_column      #find the column



print(">>>>>>>>>>>>>>>>>>>>>>>>>>>>>  Position id and name of employees who has worked for 7 consecutive days")

print("************************************************************************************")
counter=1
for i in range(2, row + 1):
    
    time_in = sheet_obj.cell(row=i, column=3)
    time_in_next = sheet_obj.cell(row=i+1, column=3)
    value1=time_in.value
    value2=time_in_next.value

    name = sheet_obj.cell(row=i, column=8)        # for the name of the employee
        
    name_next = sheet_obj.cell(row=i+1, column=8)     # for the next Employee which we will compare 
    emp_name = name.value                                 #cleared name of employee
    emp_name_next = name_next.value                       #next Employee name

    postion_id=sheet_obj.cell(row=i, column=1)
    
    
    
    
    if value1=="" or value2=="":               #for Eleminate the empty values
        counter=1
        continue
    
    else:
        date = value1.date()
        date_next=value2.date()              # here is decleae only the date from the date and time formate
        day=date.day                         # here is find the day from the date
        day_next=date_next.day                

        if day==day_next and emp_name==emp_name_next:
            continue
        elif day+1==day_next and emp_name==emp_name_next:
            counter=counter+1
            if counter==7:
                pos_id = postion_id.value
                print("ID is ->          ",pos_id)
                print("Employee name ->  ",emp_name)
        else:
            counter=1
            continue

print("************************************************************************************")
print()
print()
print(">>>>>>>>>>>>>>>>>>>>>>>>>>>>>  position id and name of employee who have less than 10 hours of time between shifts but greater than 1 hour")
print("*************************************************")

for i in range(2, row + 1):

    time_at_out = sheet_obj.cell(row=i, column=4)                       #declearing the object the of time when employee out after the first shift
    time_in_afterFirstShift = sheet_obj.cell(row=i+1, column=3)         # creting the object when employee start the second shift 

    time_Enteryafter_firstShift=time_in_afterFirstShift.value          #fine the value(time ) when employee out after the first shift
    time_out=time_at_out.value                                         #fine the value(time ) when employee start the second shift

    postion_id=sheet_obj.cell(row=i, column=1)                     #for the id of employees
    name = sheet_obj.cell(row=i, column=8)                       #for the name of employees

    if time_out=="" or time_Enteryafter_firstShift=="":           #for elemenate the empty value
        continue
    else:
        time_O=time_out.time()                             #extract only time when employee get out when they come out after first shift

        time_in_AS=time_Enteryafter_firstShift.time()     #  extract only time when employee get in for senond shift

        date_out=time_out.date()                        # find the date

        date_in= time_Enteryafter_firstShift.date()     #find the date 
        if date_out==date_in:
            shift_time_diff_hour=time_in_AS.hour - time_O.hour       # find the hour
            shift_time_diff_min =time_in_AS.minute - time_O.minute   # find the minutes
            shift_time_diff_sec =time_in_AS.second - time_O.second   # find the seconds


            total_time_diff = (shift_time_diff_hour*60*60)+(shift_time_diff_min*60)+(shift_time_diff_sec)  #here change the all value in second
            if total_time_diff>1*60*60 and total_time_diff<10*60*60:    #here change the comparesion value in sec and compare


                pos_id = postion_id.value          # id
                emp_name = name.value              # name

                print("ID is ->          ",pos_id)
                print("Employee name ->  ",emp_name)
                print("------------------------------------------------")
print("*****************************************************")

print()
print()


print(">>>>>>>>>>>>>>>>>>>>>>>>>>>>>  ID and Name of Employee Who has worked for more than 14 hours in a single shift")
print("*****************************************************")


for i in range(2, row + 1):
    timeCard = sheet_obj.cell(row=i, column=5)       # create object for the timecard
    time=timeCard.value                               

    
    if time=="":
        continue
    else:
        
        tt=time.split(":")           # I used the split function for the split the string 
        
        time_hour=int(tt[0])          # change the datatype into int for the string and also find the hour
        time_min=int(tt[1])           # change the datatype into int for the string and also find the min

        
        
        total_time_in_sec = (time_hour*60*60)+(time_min*60)    # changing all the value in the sec

        if total_time_in_sec>14*60*60:                         # compareing the value who worked 14 hour in single shift
            postion_id=sheet_obj.cell(row=i, column=1)
            name = sheet_obj.cell(row=i, column=8)
            pos_id = postion_id.value
            emp_name = name.value
            print("ID is ->          ",pos_id)
            print("Employee name ->  ",emp_name)

print("*****************************************************")

           
